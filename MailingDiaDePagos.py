from openpyxl import load_workbook
import time
import datetime
from datetime import date
from datetime import datetime
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
import os
from os import listdir
from os.path import isfile, join
import glob
import smtplib
from email.MIMEText import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.application import MIMEApplication

def SendNewMail(Usuario,contrasena,mail_destino,cuerpo,html):
    smtpObj = smtplib.SMTP('smtp.gmail.com',587)
    smtpObj.starttls()
    smtpObj.login(Usuario,contrasena)
    msg = MIMEMultipart()
    msg.attach(cuerpo)
    msg.attach(html)
    msg['From']=Usuario
    msg['Subject']="Deuda plan Uber"
    smtpObj.sendmail(Usuario,mail_destino,str(msg))
    smtpObj.quit()

def definirCorreo(ws,columna,ultima_fila):
    Semanas_a_tiempo=ws.cell(row=8,column=columna).value
    SaldoFinalAcumulado = ws.cell(row=ultima_fila,column=columna).value
    Fecha = date.today()
    if(Semanas_a_tiempo>7):#Estrella Plus
        if(SaldoFinalAcumulado>0):
            html = "Te escribo por parte de Mutuo Financiera. Para recordarte que tu renta no ha sido cubierta. Tienes un saldo de $"+ SaldoFinalAcumulado+" correspondiente al pago de tu veh&iacute;culo. Estamos seguros que se debe a un descuido, te pedimos lo pagues hoy mismo.\nSi ya ha enviado su pago, perm&iacute;tanos agradecerle.\nSi existe algo m&aacute;s en lo que podamos ayudarte no dejes de contactarnos.\n Sin m&aacute;s por el momento, seguimos a tus &oacute;rdenes."
        else:
            html = "Te escribo por parte de Mutuo Financiera. El presente tiene como motivo agradecerte tu compromiso con nosotros y hacerte saber que tu pago fue recibido de manera oportuna y puntual. Nos enorgullece saber que estas comprometido con tu proyecto y que vas al corriente en tu esquema de pagos.\n Si existe algo m$aacute;s en lo que podamos ayudarte no dejes de contactarnos. \n Sin m&aacute;s por el momento, seguimos a tus &oacute;rdenes."
    else:
        if(Semanas_a_tiempo>3):#Estrella
            if(SaldoFinalAcumulado>0):
                html = "Te escribo por parte de Mutuo Financiera donde tenemos como objetivo impulsar a M&eacute;xico a trav&eacute;s de la inclusi&oacute;n financiera de MYPYMEs como t&uacute;, por lo tanto, nos emociona estar contigo durante el proceso de adquirir tu veh&iacute;culo propio. Recuerda que si pagas a tiempo el monto de $"+ SaldoFinalAcumulado +" el d&iacute;a [Fecha adeudo], juntos podremos seguir creciendo tu negocio."
                ws.cell(row=8,column=columna).value = 0
            else:
                html = "Te escribo por parte de Mutuo Financiera. El presente tiene como motivo agradecerte tu compromiso con nosotros y hacerte saber que tu pago fue recibido de manera oportuna y puntual. Nos enorgullece saber que estas comprometido con tu proyecto y que vas al corriente en tu esquema de pagos.\nSi existe algo m&aacute;s en lo que podamos ayudarte no dejes de contactarnos.\nSin m&aacute;s por el momento, seguimos a tus &oacute;rdenes."
        else: #Busqueda Prueba o Normal
            FechaInicio = ws.cell(row=6,column=int(columna)).value
            FechaActual = date.today()
            FechaActual = datetime.combine(FechaActual,datetime.min.time())
            Diferencia_Fechas= (FechaActual-FechaInicio).days/7
            if(Diferencia_Fechas>11): #Ya no esta en periodo de prueba
                if(SaldoFinalAcumulado>0):#Debe
                    html= "Te escribo por parte de Mutuo Financiera. Desde [Fecha adeudo], tienes un atraso de XXXX en tu saldo con respecto a la renta de tu veh&iacute;culo dentro de la plataforma UBER. Solicitamos te pongas al corriente con tus pagos a la brevedad para que los intereses moratorios no sigan incrementando y podamos seguir ayud&aacute;ndote en el proceso de independizarte.\nSi existe algo m&aacute;s en lo que podamos ayudarte no dejes de contactarnos.\nSin m&aacute;s por el momento, seguimos a tus &oacute;rdenes."
                else:#No debe
                    html= "Te escribo por parte de Mutuo Financiera. Nos da mucho gusto verte crecer a trav&eacute;s de la adquisici&oacute;n de tu veh&iacute;culo propio,  por lo que nos gustar&iacute;a agradecerte que hemos recibido tu pago de manera puntual. Recuerda que al finalizar todos tus pagos, ser&aacute;s tu propio jefe.\nSi existe algo m&aacute;s en lo que podamos ayudarte no dejes de contactarnos.\nSin m&aacute;s por el momento, seguimos a tus &oacute;rdenes."
            else: #Periodo de Prueba
                if(SaldoFinalAcumulado>0):
                    html = "Te escribo por parte de Mutuo Financiera. Recuerda que los 3 primeros meses de tu contrato son un periodo de prueba para poder confiar en ti. A d&iacute;a de hoy "+ str(Fecha.day) +"/"+ str(Fecha.month)+"/"+ str(Fecha.year)+", est&aacute; pendiente el pago de $"+ str(SaldoFinalAcumulado)+". Te solicitamos liquides tu saldo a la brevedad posible, de lo contrario nos veremos obligados a terminar con el periodo de prueba y nuestra relaci&oacute;n contractual."
                else:
                    html = "Te escribo por parte de Mutuo Financiera. El presente tiene como motivo agradecerte tu compromiso con nosotros y hacerte saber que tu pago fue recibido de manera oportuna y puntual. Nos enorgullece saber que estas comprometido con tu proyecto y que vas al corriente en tu esquema de pagos.\n Si existe algo m$aacute;s en lo que podamos ayudarte no dejes de contactarnos. \n Sin m&aacute;s por el momento, seguimos a tus &oacute;rdenes."



    return html

def ComponerMail(html):
    """<!DOCTYPE html>
    <html lang="en" xmlns="http://www.w3.org/1999/xhtml" xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office">
    <head>
        <meta charset="utf-8"> <!-- utf-8 works for most cases -->
        <meta name="viewport" content="width=device-width"> <!-- Forcing initial-scale shouldn't be necessary -->
        <meta http-equiv="X-UA-Compatible" content="IE=edge"> <!-- Use the latest (edge) version of IE rendering engine -->
        <meta name="x-apple-disable-message-reformatting">  <!-- Disable auto-scale in iOS 10 Mail entirely -->
        <title></title> <!-- The title tag shows in email notifications, like Android 4.4. -->

        <!-- Web Font / @font-face : BEGIN -->
        <!-- NOTE: If web fonts are not required, lines 10 - 27 can be safely removed. -->

        <!-- Desktop Outlook chokes on web font references and defaults to Times New Roman, so we force a safe fallback font. -->
        <!--[if mso]>
            <style>
                * {
                    font-family: sans-serif !important;
                }
            </style>
        <![endif]-->

        <!-- All other clients get the webfont reference; some will render the font and others will silently fail to the fallbacks. More on that here: http://stylecampaign.com/blog/2015/02/webfont-support-in-email/ -->
        <!--[if !mso]><!-->
        <!-- insert web font reference, eg: <link href='https://fonts.googleapis.com/css?family=Roboto:400,700' rel='stylesheet' type='text/css'> -->
        <!--<![endif]-->

        <!-- Web Font / @font-face : END -->

        <!-- CSS Reset : BEGIN -->
        <link href="https://fonts.googleapis.com/css?family=Lato|Nunito|Roboto" rel="stylesheet">
        <style>
            /* What it does: Remove spaces around the email design added by some email clients. */
            /* Beware: It can remove the padding / margin and add a background color to the compose a reply window. */
            html,
            body {
                margin: 0 auto !important;
                padding: 0 !important;
                height: 100% !important;
                width: 100% !important;
                font-family: 'Nunito', sans-serif !important;
            }

            /* What it does: Stops email clients resizing small text. */
            * {
                -ms-text-size-adjust: 100%;
                -webkit-text-size-adjust: 100%;
                font-family: 'Nunito', sans-serif !important;
            }

            /* What it does: Centers email on Android 4.4 */
            div[style*="margin: 16px 0"] {
                margin: 0 !important;
                font-family: 'Nunito', sans-serif !important;
            }

            /* What it does: Stops Outlook from adding extra spacing to tables. */
            table,
            td {
                mso-table-lspace: 0pt !important;
                mso-table-rspace: 0pt !important;
            }

            /* What it does: Fixes webkit padding issue. Fix for Yahoo mail table alignment bug. Applies table-layout to the first 2 tables then removes for anything nested deeper. */
            table {
                border-spacing: 0 !important;
                border-collapse: collapse !important;
                table-layout: fixed !important;
                margin: 0 auto !important;
                font-family: 'Nunito', sans-serif !important;
            }
            table table table {
                table-layout: auto;
            }

            /* What it does: Uses a better rendering method when resizing images in IE. */
            img {
                -ms-interpolation-mode:bicubic;
            }

            /* What it does: A work-around for email clients meddling in triggered links. */
            *[x-apple-data-detectors],  /* iOS */
            .x-gmail-data-detectors,    /* Gmail */
            .x-gmail-data-detectors *,
            .aBn {
                border-bottom: 0 !important;
                cursor: default !important;
                color: inherit !important;
                text-decoration: none !important;
                font-size: inherit !important;
                font-family: 'Nunito', sans-serif !important;
                font-weight: inherit !important;
                line-height: inherit !important;
            }

            /* What it does: Prevents Gmail from displaying an download button on large, non-linked images. */
            .a6S {
                display: none !important;
                opacity: 0.01 !important;
            }
            /* If the above doesn't work, add a .g-img class to any image in question. */
            img.g-img + div {
                display: none !important;
            }

            /* What it does: Prevents underlining the button text in Windows 10 */
            .button-link {
                text-decoration: none !important;
            }

            /* What it does: Removes right gutter in Gmail iOS app: https://github.com/TedGoas/Cerberus/issues/89  */
            /* Create one of these media queries for each additional viewport size you'd like to fix */
            /* Thanks to Eric Lepetit @ericlepetitsf) for help troubleshooting */
            @media only screen and (min-device-width: 375px) and (max-device-width: 413px) { /* iPhone 6 and 6+ */
                .email-container {
                    min-width: 375px !important;
                }
            }

            /* What it does: Forces Gmail app to display email full width */
            div > u ~ div .gmail {
                min-width: 100vw;
            }

        </style>
        <!-- CSS Reset : END -->

        <!-- Progressive Enhancements : BEGIN -->
        <style>

        /* What it does: Hover styles for buttons */
        .button-td,
        .button-a {
            transition: all 100ms ease-in;
        }
        .button-td:hover,
        .button-a:hover {
            background: #555555 !important;
            border-color: #555555 !important;
        }

        /* Media Queries */
        @media screen and (max-width: 480px) {

            /* What it does: Forces elements to resize to the full width of their container. Useful for resizing images beyond their max-width. */
            .fluid {
                width: 100% !important;
                max-width: 100% !important;
                height: auto !important;
                margin-left: auto !important;
                margin-right: auto !important;
            }

            /* What it does: Forces table cells into full-width rows. */
            .stack-column,
            .stack-column-center {
                display: block !important;
                width: 100% !important;
                max-width: 100% !important;
                direction: ltr !important;
            }
            /* And center justify these ones. */
            .stack-column-center {
                text-align: center !important;
            }

            /* What it does: Generic utility class for centering. Useful for images, buttons, and nested tables. */
            .center-on-narrow {
                text-align: center !important;
                display: block !important;
                margin-left: auto !important;
                margin-right: auto !important;
                float: none !important;
            }
            table.center-on-narrow {
                display: inline-block !important;
            }

            /* What it does: Adjust typography on small screens to improve readability */
            .email-container p {
                font-size: 17px !important;
            }
        }

        </style>

        <!-- Progressive Enhancements : END -->

        <!-- What it does: Makes background images in 72ppi Outlook render at correct size. -->
        <!--[if gte mso 9]>
        <xml>
            <o:OfficeDocumentSettings>
                <o:AllowPNG/>
                <o:PixelsPerInch>96</o:PixelsPerInch>
            </o:OfficeDocumentSettings>
        </xml>
        <![endif]-->

    </head>
    <body width="100%" bgcolor="#fafafa" style="margin: 0; mso-line-height-rule: exactly;">
        <center style="width: 100%; background: #fafafa; text-align: left;">

            <!-- Visually Hidden Preheader Text : BEGIN -->
            <div style="display: none; font-size: 1px; line-height: 1px; max-height: 0px; max-width: 0px; opacity: 0; overflow: hidden; mso-hide: all; font-family: sans-serif;">
                (Optional) This text will appear in the inbox preview, but not the email body. It can be used to supplement the email subject line or even summarize the email's contents. Extended text preheaders (~490 characters) seems like a better UX for anyone using a screenreader or voice-command apps like Siri to dictate the contents of an email. If this text is not included, email clients will automatically populate it using the text (including image alt text) at the start of the email's body.
            </div>
            <!-- Visually Hidden Preheader Text : END -->

            <!--
                Set the email width. Defined in two places:
                1. max-width for all clients except Desktop Windows Outlook, allowing the email to squish on narrow but never go wider than 680px.
                2. MSO tags for Desktop Windows Outlook enforce a 680px width.
                Note: The Fluid and Responsive templates have a different width (600px). The hybrid grid is more "fragile", and I've found that 680px is a good width. Change with caution.
            -->
            <div style="max-width: 680px; margin: auto;" class="email-container">
                <!--[if mso]>
                <table role="presentation" cellspacing="0" cellpadding="0" border="0" width="680" align="center">
                <tr>
                <td>
                <![endif]-->

                <!-- Email Body : BEGIN -->
                <table role="presentation" cellspacing="0" cellpadding="0" border="0" align="center" width="100%" height="670px" style="max-width: 680px;" class="email-container">

                    <!-- 1 Column Text + Button : BEGIN -->
                    <tr>
                        <td bgcolor="#fafafa">
                            <table role="presentation" cellspacing="0" cellpdding="0" border="0" width="100%">
                                <tr>
                                    <img src="http://i347.photobucket.com/albums/p444/Gupie21/grfico_web-02_zpsiymaqryr.png" alt="alt_text" border="0" align="left" style="position: relative; width: 100%; height: auto;">
                                </tr>
                                <tr>
                                    <td style="padding: 0px 40px 50px; text-align: left;">
                                        <span><h1 style="margin: 0; font-family: sans-serif; font-size: 18px; line-height: 125%; color: #01558B; font-weight: normal; padding-top: 5rem">01/01/2018</h1></span>
                                    </td>
                                </tr>
                                <tr>
                                    <td style="padding: 0px 40px 20px; text-align: left;">
                                        <h1 style="margin: 0.1rem 0px 0px; margin-bottom: 50px; font-family: sans-serif; font-size: 18px; line-height: 125%; color: #01558B; font-weight: normal;">Estimad@ """+ Nombre+"""</h1>
                                    </td>
                                </tr>
                                <tr>
                                    <td style="padding: 0 40px 50px; font-family: sans-serif; font-size: 15px; line-height: 140%; color: #01558b; text-align: justify;">
                                        <p style="margin: 0; margin-bottom: 25px;">"""+html+"""</p>
                                    </td>
                                </tr>
                                <tr>
                                    <td style="padding: 0 40px 40px; font-family: sans-serif; font-size: 15px; line-height: 140%; color: #555555;">
                                        <!-- Button : BEGIN -->
                                        <table role="presentation" cellspacing="0" cellpadding="0" border="0" align="center" style="margin: auto;">
                                            <tr>
                                                <td style="padding: 0px 0; text-align: center">
                                                    <img src="http://i347.photobucket.com/albums/p444/Gupie21/correo2_zpsutovvifm.jpg" width="150" height="50" alt="alt_text" border="0" style="height: auto; background: #dddddd; font-family: sans-serif; font-size: 15px; line-height: 140%; color: #555555;">
                                                </td>
                                            </tr>
                                        </table>
                                        <!-- Button : END -->
                                    </td>
                                </tr>
                                <tr>
                                    <td style="padding: 0px 40px 80px; text-align: center;">
                                        <h1 style="margin: 0; font-family: sans-serif; font-size: 18px; line-height: 125%; color: #01558B; font-weight: bold;">#JuntosPodemosMás</h1>
                                    </td>
                                </tr>
                                <tr>
                                    <td style="padding: 0px 40px 80px; text-align: center;">
                                        <span style="margin-right: 2%; color: #01558B;">Contacto: <a style="text-decoration: none;" href="tel:+5550866576" title="">[55] 50 86 65 76</a></span>
                                        <span style="color: #01558B;"><a href="www.mutuofinanciera.com" title="" style="text-decoration: none;">www.mutuofinanciera.com</a></span>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                    <!-- 1 Column Text + Button : END -->
                </table>
                <!-- Email Body : END -->

                <!--[if mso]>
                </td>
                </tr>
                </table>
                <![endif]-->
            </div>

        </center>
    </body>
    </html>
"""

UberFile = [f for f in listdir('C:\Users\Mutuo Midgard\Box Sync\ISA F\UBER\Pagos Uber') if isfile(join('C:\Users\Mutuo Midgard\Box Sync\ISA F\UBER\Pagos Uber', f))]
print UberFile



wb=load_workbook('C:\Users\Mutuo Midgard\Box Sync\ISA F\UBER\Pagos Uber/'+str(UberFile[0]), data_only=True)
ws=wb.active
ultima_fila = ws.max_row
ultima_columna = ws.max_column
Usuario = 'lserio@mutuofinanciera.com'
contrasena = 'Fhvy7032'
print(ultima_columna)

for x in range(3,ultima_columna):
    if(True):
        Nombre = ws.cell(row=2,column=x).value
        #Componer html
        html= definirCorreo(ws,x,ultima_fila)
        cuerpo= ComponerMail(Nombre, html)
        print(html)
        print('Numero: '+str(x))
        print(Nombre)
        #Extraer mail de destino
        #SendNewMail(Usuario,contrasena,mail_destino,cuerpo,html)
