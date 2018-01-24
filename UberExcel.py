import sys
reload(sys)
sys.setdefaultencoding('utf-8')
import os
from os import listdir
from os.path import isfile, join
import glob
import csv
from xlsxwriter.workbook import Workbook
import smtplib
import time
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

def LetrasExcel(Numero):
    return {1:'A',2:'B',3:'C',4:'D',5:'E',6:'F',7:'G',8:'H',9:'I',10:'J',11:'K',12:'L',13:'M',14:'N',15:'O',16:'P',17:'Q',18:'R'}[Numero]

def  Formato(ws,ultima_fila):
    #Definimos los colores del formato previamente establecido por Isa F
    FontDate = Font(bold=True,size=11,name='Calibri',color='FFFFFF')
    FontStandard = Font(name='Calibri',color='FFFFFF')
    GreyBackground = PatternFill(fill_type='solid',start_color='CCCCCC',end_color='CCCCCC')
    Results = Font(bold=True)
    BlueBackground = PatternFill(fill_type='solid',start_color='01558b',end_color='01558b')
    #Definimos donde van a ser aplicados estos estilos
    FilaDate = ultima_fila+2
    Date = 'B' + str(FilaDate)
    DateCell=ws[Date]
    FilaRetencion = ultima_fila+3
    FilaComplementoPagado = ultima_fila+5
    FilaResultados = ultima_fila+6
    FilaSaldoFinal = ultima_fila+7
    FilaSaldoFinalAcumulado = ultima_fila+8
    SaldoFinalAcumulado = 'B' + str(FilaSaldoFinalAcumulado)
    SaldoFinalAcumuladoCell = ws[SaldoFinalAcumulado]
    #Aplicamos estilos
    #RangoTotal = 'Currency'
    LimiteColumnas = ws.max_column
    for x in range(3,LimiteColumnas):
        Letra = LetrasExcel(x)
        CoordDate = Letra+str(FilaDate)
        CeldaDate=[CoordDate]
        CoordRetencion = Letra+str(FilaRetencion)
        CeldaRetencion=ws[CoordRetencion]
        CeldaRetencion.fill=GreyBackground
        CoordFaltante=Letra+str(FilaRetencion)
        CeldaFaltante=ws[CoordFaltante]
        CoordComplemento = Letra+str(FilaComplementoPagado)
        CeldaComplemento=ws[CoordComplemento]
        CeldaComplemento.fill=GreyBackground
        CoordResultados = Letra+str(FilaResultados)
        CeldaResultados=ws[CoordResultados]
        CoordSaldoFinal = Letra+str(FilaResultados)
        CeldaSaldoFinal=ws[CoordSaldoFinal]
        CeldaSaldoFinal.font=Results
        #ws.cell(row=FilaRetencion,column=x).fill = GreyBackground
        #ws.cell(row=FilaComplementoPagado,column=x).fill = GreyBackground
        #ws.cell(row=FilaSaldoFinal,column=x).font=Results
    SaldoFinalAcumuladoCell.font = Results
    DateCell.font=FontDate
    DateCell.fill = BlueBackground

def CalculoFecha(ws,ultima_fila):
    Ultima_fecha = ws.cell(row=ultima_fila-5,column=2).value
    print(type(Ultima_fecha))
    semana = datetime.timedelta(days=7)
    Fecha_Final = Ultima_fecha+semana
    print(Fecha_Final)
    FilaDate = ultima_fila+2
    Date = 'B' + str(FilaDate)
    DateCell=ws[Date]
    ws[Date]=Fecha_Final
    FilaRetencion = ultima_fila+3
    Retencion = 'B' + str(FilaRetencion)
    ws[Retencion]='Retencion'
    FilaFaltante = ultima_fila+4
    Faltante = 'B' + str(FilaFaltante)
    ws[Faltante]='Faltante por pagar'
    FilaComplemento = ultima_fila+5
    Complemento = 'B' + str(FilaComplemento)
    ws[Complemento]='Complemento pagado'
    FilaTotal = ultima_fila+6
    Total = 'B' + str(FilaTotal)
    ws[Total]='Total Pagado'
    FilaSaldoFinal = ultima_fila+7
    SaldoFinal = 'B' + str(FilaSaldoFinal)
    ws[SaldoFinal]='SaldoFinalAcumulado'
    ws[Date].number_format = 'd-mmm'
    return(Fecha_Final)

def BusquedaID(id_unico,Pago_cargado,ws,ultima_fila,Fecha_Final):
    for x in range(3,15):
        IdentificadorExcel = ws.cell(row=3,column=x).value
        if(IdentificadorExcel == id_unico):
            Producto = ws.cell(row=7,column=x).value
            print('Llegue')
            print(type(str(Producto)))
            if(str(Producto)=='S'):
                Sedan(Pago_cargado,ws,ultima_fila,Fecha_Final,x)
                print('SiEntre')
            else:
                if(str(Producto)=='V'):
                    Versa(Pago_cargado,ws,ultima_fila,Fecha_Final,x)
                else:
                    if(str(Producto)=='SE'):
                        SinEnganche(Pago_cargado,ws,ultima_fila,Fecha_Final,x)
                    else:
                        if(str(Producto)=='ME'):
                            MedioEnganche(Pago_cargado,ws,ultima_fila,Fecha_Final,x)
                        else:
                            if(str(Producto)=='JM'):
                                JorgeMaldonado(Pago_cargado,ws,ultima_fila,Fecha_Final,x)


def Sedan(Pago_cargado,ws,ultima_fila,Fecha_Final,columna):
    FechaInicio = ws.cell(row=6,column=int(columna)).value
    FechaActual = ws.cell(row=ultima_fila+2,column=2).value
    Diferencia_Fechas= (FechaActual-FechaInicio).days/7
    if(Diferencia_Fechas<209):
        LetraCliente=LetrasExcel(int(columna))
        PagoCoord = LetraCliente+str(ultima_fila+3)
        print(PagoCoord)
        ws[PagoCoord]=float(Pago_cargado)
        Deuda = ws.cell(row=ultima_fila,column=int(columna)).value
        CoordDeuda = LetraCliente+str(ultima_fila+2)
        DeudaFinal = float(Deuda)+2730.27
        ws[CoordDeuda] = DeudaFinal
        #Desde aqui -- Lo que falta por pagar
        CoordFaltante = LetraCliente+str(ultima_fila+4)
        ws[CoordFaltante] = DeudaFinal - float(ws.cell(row=ultima_fila+3,column=columna).value)
        #Formulas, faltante
        ws[LetraCliente+str(ultima_fila+6)] = "="+PagoCoord+"+"+LetraCliente+str(ultima_fila+5)
        #Saldo Final Acumulado
        ws[LetraCliente+str(ultima_fila+7)] = "="+CoordDeuda+"-"+LetraCliente+str(ultima_fila+6)
        #Formato Numero
        ws[PagoCoord].number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws[CoordDeuda].number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws[CoordFaltante].number_format='"$"#,##0.00_);("$"#,##0.00)'
        ws[LetraCliente+str(ultima_fila+5)].number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws[LetraCliente+str(ultima_fila+6)].number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws[LetraCliente+str(ultima_fila+7)].number_format = '[Red]"$"#,##0.00_);[Color 10]"-$"#,##0.00'




def Versa(Pago_cargado,ws,ultima_fila,Fecha_Final,columna):
    FechaInicio = ws.cell(row=6,column=int(columna)).value
    FechaActual = ws.cell(row=ultima_fila+2,column=2).value
    Diferencia_Fechas= (FechaActual-FechaInicio).days/7
    if(Diferencia_Fechas<209):
        LetraCliente=LetrasExcel(int(columna))
        PagoCoord = LetraCliente+str(ultima_fila+3)
        print(PagoCoord)

        ws[PagoCoord]=float(Pago_cargado)

        Deuda = ws.cell(row=ultima_fila,column=int(columna)).value
        CoordDeuda = LetraCliente+str(ultima_fila+2)
        DeudaFinal = float(Deuda)+1860
        ws[CoordDeuda] = DeudaFinal
        #Desde aqui -- Lo que falta por pagar
        CoordFaltante = LetraCliente+str(ultima_fila+4)
        ws[CoordFaltante] = DeudaFinal - float(ws.cell(row=ultima_fila+3,column=columna).value)
        #Formulas, faltante
        ws[LetraCliente+str(ultima_fila+6)] = "="+PagoCoord+"+"+LetraCliente+str(ultima_fila+5)
        #Saldo Final Acumulado
        ws[LetraCliente+str(ultima_fila+7)] = "="+CoordDeuda+"-"+LetraCliente+str(ultima_fila+6)
        #Formato Numero
        ws[PagoCoord].number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws[CoordDeuda].number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws[CoordFaltante].number_format='"$"#,##0.00_);("$"#,##0.00)'
        ws[LetraCliente+str(ultima_fila+5)].number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws[LetraCliente+str(ultima_fila+6)].number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws[LetraCliente+str(ultima_fila+7)].number_format = '[Red]"$"#,##0.00_);[Color 10]"-$"#,##0.00'



def SinEnganche(Pago_cargado,ws,ultima_fila,Fecha_Final,columna):
    FechaInicio = ws.cell(row=6,column=int(columna)).value
    FechaActual = ws.cell(row=ultima_fila+2,column=2).value
    Diferencia_Fechas= (FechaActual-FechaInicio).days/7
    if(Diferencia_Fechas<217):
        LetraCliente=LetrasExcel(int(columna))
        PagoCoord = LetraCliente+str(ultima_fila+3)
        print(PagoCoord)
        ws[PagoCoord]=float(Pago_cargado)
        Deuda = ws.cell(row=ultima_fila,column=int(columna)).value
        CoordDeuda = LetraCliente+str(ultima_fila+2)
        print(Deuda)
        if(Deuda!='#VALUE!'):
            if(Diferencia_Fechas<25):
                DeudaFinal = float(Deuda)+2500
                ws[CoordDeuda] = DeudaFinal
            else:
                DeudaFinal= float(Deuda+2000)
                ws[CoordDeuda] = DeudaFinal
            #Desde aqui -- Lo que falta por pagar
            CoordFaltante = LetraCliente+str(ultima_fila+4)
            ws[CoordFaltante] = DeudaFinal - float(ws.cell(row=ultima_fila+3,column=columna).value)
        #Formulas, faltante
        ws[LetraCliente+str(ultima_fila+6)] = "="+PagoCoord+"+"+LetraCliente+str(ultima_fila+5)
        #Saldo Final Acumulado
        ws[LetraCliente+str(ultima_fila+7)] = "="+CoordDeuda+"-"+LetraCliente+str(ultima_fila+6)
        #Formato Numero
        ws[PagoCoord].number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws[CoordDeuda].number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws[CoordFaltante].number_format='"$"#,##0.00_);("$"#,##0.00)'
        ws[LetraCliente+str(ultima_fila+5)].number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws[LetraCliente+str(ultima_fila+6)].number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws[LetraCliente+str(ultima_fila+7)].number_format = '[Red]"$"#,##0.00_);[Color 10]"-$"#,##0.00'

def MedioEnganche(Pago_cargado,ws,ultima_fila,Fecha_Final,columna):
    FechaInicio = ws.cell(row=6,column=columna).value
    FechaActual = ws.cell(row=ultima_fila+2,column=2).value
    Diferencia_Fechas= (FechaActual-FechaInicio).days/7
    if(Diferencia_Fechas<209):
        LetraCliente=LetrasExcel(int(columna))
        PagoCoord = LetraCliente+str(ultima_fila+3)
        print(PagoCoord)
        ws[PagoCoord]=float(Pago_cargado)
        Deuda = ws.cell(row=ultima_fila,column=int(columna)).value
        CoordDeuda = LetraCliente+str(ultima_fila+2)
        if(Diferencia_Fechas<25):
            DeudaFinal = float(Deuda)+2500
            ws[CoordDeuda] = DeudaFinal
        else:
            DeudaFinal= float(Deuda+2000)
            ws[CoordDeuda] = DeudaFinal
        #Desde aqui -- Lo que falta por pagar
        CoordFaltante = LetraCliente+str(ultima_fila+4)
        ws[CoordFaltante] = DeudaFinal - float(ws.cell(row=ultima_fila+3,column=columna).value)
        #Formulas, faltante
        ws[LetraCliente+str(ultima_fila+6)] = "="+PagoCoord+"+"+LetraCliente+str(ultima_fila+5)
        #Saldo Final Acumulado
        ws[LetraCliente+str(ultima_fila+7)] = "="+CoordDeuda+"-"+LetraCliente+str(ultima_fila+6)
        #Formato Numero
        ws[PagoCoord].number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws[CoordDeuda].number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws[CoordFaltante].number_format='"$"#,##0.00_);("$"#,##0.00)'
        ws[LetraCliente+str(ultima_fila+5)].number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws[LetraCliente+str(ultima_fila+6)].number_format = '"$"#,##0.00_);("$"#,##0.00)'
        ws[LetraCliente+str(ultima_fila+7)].number_format = '[Red]"$"#,##0.00_);[Color 10]"-$"#,##0.00'


def JorgeMaldonado(Pago_cargado,ws,ultima_fila,Fecha_Final,columna):
    FechaInicio = ws.cell(row=6,column=columna).value
    FechaActual = ws.cell(row=ultima_fila+2,column=2).value
    Diferencia_Fechas= (FechaActual-FechaInicio).days/7
    print('Fechas hasta hoy:'+str(Diferencia_Fechas))
    if(Diferencia_Fechas<16):
        LetraCliente=LetrasExcel(int(columna))
        PagoCoord = LetraCliente+str(ultima_fila+3)
        print(PagoCoord)
        ws[PagoCoord]=float(Pago_cargado)
        Deuda = ws.cell(row=ultima_fila,column=int(columna)).value
        CoordDeuda = LetraCliente+str(ultima_fila+2)
        DeudaFinal = float(Deuda)+3640.36
        ws[CoordDeuda] = DeudaFinal
        #Desde aqui -- Lo que falta por pagar
        CoordFaltante = LetraCliente+str(ultima_fila+4)
        ws[CoordFaltante] = DeudaFinal - float(ws.cell(row=ultima_fila+3,column=columna).value)
    else:
        LetraCliente=LetrasExcel(int(columna))
        PagoCoord = LetraCliente+str(ultima_fila+3)
        print(PagoCoord)
        ws[PagoCoord]=float(Pago_cargado)
        Deuda = ws.cell(row=ultima_fila,column=int(columna)).value
        CoordDeuda = LetraCliente+str(ultima_fila+2)
        DeudaFinal = float(Deuda)+3640.36
        ws[CoordDeuda] = DeudaFinal
        #Desde aqui -- Lo que falta por pagar
        CoordFaltante = LetraCliente+str(ultima_fila+4)
        ws[CoordFaltante] = DeudaFinal - float(ws.cell(row=ultima_fila+3,column=columna).value)
        CoordProducto=LetraCliente+'7'
        ws[CoordProducto]= 'S'
    #Formulas, faltante
    ws[LetraCliente+str(ultima_fila+6)] = "="+PagoCoord+"+"+LetraCliente+str(ultima_fila+5)
    #Saldo Final Acumulado
    ws[LetraCliente+str(ultima_fila+7)] = "="+CoordDeuda+"-"+LetraCliente+str(ultima_fila+6)
    #Formato Numero
    ws[PagoCoord].number_format = '"$"#,##0.00_);("$"#,##0.00)'
    ws[CoordDeuda].number_format = '"$"#,##0.00_);("$"#,##0.00)'
    ws[CoordFaltante].number_format='"$"#,##0.00_);("$"#,##0.00)'
    ws[LetraCliente+str(ultima_fila+5)].number_format = '"$"#,##0.00_);("$"#,##0.00)'
    ws[LetraCliente+str(ultima_fila+6)].number_format = '"$"#,##0.00_);("$"#,##0.00)'
    ws[LetraCliente+str(ultima_fila+7)].number_format = '[Red]"$"#,##0.00_);[Color 10]"-$"#,##0.00'



UberFile = [f for f in listdir('C:\Users\Mutuo Midgard\Box Sync\ISA F\UBER\Pagos Uber') if isfile(join('C:\Users\Mutuo Midgard\Box Sync\ISA F\UBER\Pagos Uber', f))]
print UberFile



wb=load_workbook('C:\Users\Mutuo Midgard\Box Sync\ISA F\UBER\Pagos Uber/'+str(UberFile[0]), data_only=True)
ws = wb.active
ultima_fila = ws.max_row
print(ultima_fila)

execfile('csvtoexcel.py')

VisorFile = [f for f in listdir('C:\Users\Mutuo Midgard\Box Sync\ISA F\UBER\Pagos Uber\PagosArkafin') if isfile(join('C:\Users\Mutuo Midgard\Box Sync\ISA F\UBER\Pagos Uber\PagosArkafin', f))]

Formato(ws,ultima_fila)
Fecha_Final=CalculoFecha(ws,ultima_fila)
wb2=load_workbook('C:\Users\Mutuo Midgard\Box Sync\ISA F\UBER\Pagos Uber\PagosArkafin/'+VisorFile[0], data_only=True)
UberToArkafin=wb2.active
Ultima_Fila_Arkafin=UberToArkafin.max_row
for x in range(2,Ultima_Fila_Arkafin+1):
    id_unico = UberToArkafin.cell(row=x,column=2).value
    Pago_cargado = UberToArkafin.cell(row=x,column=4).value
    BusquedaID(id_unico,Pago_cargado,ws,ultima_fila,Fecha_Final)



mesnombre='%02d' % datetime.date.today().month
print(mesnombre)
dianombre='%02d' % datetime.date.today().day
print(dianombre)
year = datetime.date.today().year
print(year)
yearstr = str(year)
yearnom = yearstr[2] + yearstr[3]
print(yearnom)
wb.save('C:\Users\Mutuo Midgard\Box Sync\ISA F\UBER\Pagos Uber/'+yearnom+mesnombre+dianombre+' Pagos Uber.xlsx')
os.rename('C:\Users\Mutuo Midgard\Box Sync\ISA F\UBER\Pagos Uber/'+str(UberFile[0]),'C:\Users\Mutuo Midgard\Box Sync\ISA F\UBER\Pagos Uber/Anteriores/'+str(UberFile[0]))
os.remove('C:\Users\Mutuo Midgard\Box Sync\ISA F\UBER\Pagos Uber\PagosArkafin/'+VisorFile[0])
